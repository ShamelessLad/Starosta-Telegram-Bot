# bot.py
# Main source code of the bot
# Copyright 2021 by Mikhail Korobkov, Kamil Muradov

import logging, random, datetime, pytz, os, requests, openpyxl, time, sys
from openpyxl.utils.indexed_list import IndexedList

from collections.abc import Iterable
from bs4 import BeautifulSoup
from telegram.ext.dispatcher import Dispatcher
from classes import *

from telegram import (
    Poll,
    ParseMode,
    KeyboardButton,
    KeyboardButtonPollType,
    ReplyKeyboardMarkup,
    ReplyKeyboardRemove,
    Update,
    InputMediaDocument
)
from telegram.ext import (
    Updater,
    CommandHandler,
    PollAnswerHandler,
    PollHandler,
    MessageHandler,
    Filters,
    CallbackContext,
)

logging.basicConfig (
    stream=sys.stdout,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', level=logging.INFO
)
logger = logging.getLogger(__name__)

def helper (update: Update, context: CallbackContext) -> None:
    update.message.reply_html (
        "Здравствуй, я бот староста! Список команд:\n"
        "<i>/toggle_email</i> - включение/отключение рассылки новых сообщений с почты группы в личные сообщения\n"
        "<i>/enable_schedule &lt;минуты&gt;</i> - включение рассылки расписания на завтра после конца последней пары в личные сообщения. Минуты: 0-269 (0-4.5 часа)\n"
        "<i>/disable_schedule</i> - отключение рассылки расписания\n"
        "<i>/id</i> - получение id своего профиля\n"
        "<i>/help</i> - вывод списка команд бота\n"
    )

def poll(context: CallbackContext) -> None:
    """Отправляет опрос присутствия на паре"""

    job = context.job
    questions = ["На парах!", "Отдыхаю!"]
    message = context.bot.send_poll(
            job.context, 
            "Ты сегодня на паре?", 
            questions, 
            is_anonymous=False, 
            allows_multiple_answers=False,
    )
    payload = {
        "attendance": {
            "absent": [],
            "not_absent": [],
            "no_vote": [],
        },
        "poll_id": message.poll.id,
        "questions": questions,
        "message_id": message.message_id,
        "chat_id": job.context,
        "answers": 0,
    }   

    if "poll" not in context.bot_data.keys():
        context.bot_data["poll"] = {}
    context.bot_data["poll"].update(payload)
    print(f"ADDING JOB CLOSE_POLL, WILL BE EXECUTED AT")
    print(context.job_queue.run_once (
            callback=close_poll, 
            when=7200, 
            context=payload,
    ).next_t)

def update_student_vote(student_id: int, option_id: int, context: CallbackContext) -> None:
    """
    Обновляем голос студента в bot_data
    Если он не голосовал, его id просто добавится в attendance
    Если переголосовал, его голос перезапишется в новый вариант ответа
    """

    for attend_st, lst in context.bot_data["poll"]["attendance"].items():
        if student_id in lst:
            context.bot_data["poll"]["attendance"][attend_st].remove(student_id)
    attendance_key = ["not_absent", "absent"][option_id]
    context.bot_data["poll"]["attendance"][attendance_key].append(student_id)

def fill_not_voted(context: CallbackContext) -> None:
    """
    Заполняет словарь в bot_data непроголосовавшими
    """
    voted_students = set(context.bot_data["poll"]["attendance"]["not_absent"]).union(set(context.bot_data["poll"]["attendance"]["absent"]))
    not_voted_students = set(context.bot_data["students_names_by_ids"].keys()).difference(voted_students)
    context.bot_data["poll"]["attendance"]["not_voted"] = list(not_voted_students)

def send_attendance_report(context: CallbackContext) -> None:
    fill_not_voted(context)
    context.bot.send_message(
        chat_id = context.bot_data["starosta_id"],
        text    = "Будут:\n    {not_absent}\n\nНе будут:\n    {absent}\n\nНе голосовали:\n    {not_voted}".format(
            not_absent  = "\n    ".join([f"{index+1}) "+context.bot_data["students_names_by_ids"][student_id] for index, student_id in enumerate(context.bot_data["poll"]["attendance"]["not_absent"])]),
            absent      = "\n    ".join([f"{index+1}) "+context.bot_data["students_names_by_ids"][student_id] for index, student_id in enumerate(context.bot_data["poll"]["attendance"]["absent"])]),
            not_voted   = "\n    ".join([f"{index+1}) "+context.bot_data["students_names_by_ids"][student_id] for index, student_id in enumerate(context.bot_data["poll"]["attendance"]["not_voted"])])
        )
    )

def receive_poll_answer(update: Update, context: CallbackContext) -> None:
    """получение результатов опроса и зего закрытие по достижении кол-ва голосов membercount-1"""

    answer = update.poll_answer
    #poll_id = answer.poll_id
    student_id = update.effective_user.id
    chat_id = context.bot_data["poll"]["chat_id"]
    membercount = context.bot.get_chat_members_count (chat_id)

    if student_id not in context.bot_data["students_names_by_ids"].keys():
        context.bot.send_message (
            chat_id=chat_id,
            text="Женя, ливни с группы."
        )
        return

    try:
        questions = context.bot_data["poll"]["questions"]
    except KeyError:
        return

    selected_options = answer.option_ids
    
    for question_id in selected_options:
        context.bot_data["poll"]["answers"] += 1
        update_student_vote(student_id, question_id, context)

    if context.bot_data["poll"]["answers"] == membercount - 2: #минус бот и куратор
        remove_job("close_poll", context)
        context.bot.stop_poll(
            context.bot_data["poll"]["chat_id"], 
            context.bot_data["poll"]["message_id"]
        )
        send_attendance_report(context)
        context.bot.delete_message (
            chat_id=context.bot_data["poll"]["chat_id"], 
            message_id=context.bot_data["poll"]["message_id"]
        ) 
        context.bot_data.pop("poll", None)

def close_poll (context: CallbackContext) -> None: 
    """Процедура для закрытия опроса по прошествии времени"""

    chat_id=context.job.context["chat_id"]
    message_id=context.job.context["message_id"]
    context.bot.stop_poll (chat_id, message_id)
    send_attendance_report(context)
    context.bot.delete_message (
            chat_id=context.bot_data["poll"]["chat_id"], 
            message_id=context.bot_data["poll"]["message_id"]
    ) 
    context.bot_data.pop("poll", None)

def remove_job (name, context: CallbackContext):
    """Функция для удаления таймеров"""

    jobs = context.job_queue.get_jobs_by_name(name)
    
    if not jobs:
        return False
    for job in jobs:
        job.schedule_removal()
    return True

def id_and_count (update: Update, context: CallbackContext) -> None:
    """id чата и количество участников в нем"""

    if update.effective_user.id not in get_admins(update) and update.effective_chat.type != "private":
        update.message.reply_text("b-baka~")
        return
    chat_id = update.message.chat.id
    membercount = context.bot.get_chat_members_count (chat_id)
    update.message.reply_text (
        #f"Количество участников: {membercount}\n"
        f"Id: {chat_id}" 
    )
    return (chat_id, membercount)
    
def get_latest_emails_job_callback(context: CallbackContext) -> None:
    chat_id = context.job.context
    messages_list = context.bot_data["emailgetter"].get_newer_messages(context.bot_data["last_email_id"])
    if messages_list:
        context.bot_data["last_email_id"] = messages_list[0]["id"]
    for msg in messages_list:
        attachments = [InputMediaDocument(media=open(attachment_dir, "rb")) for attachment_dir in msg["attachments_dirs"]]
        if attachments:
            attachments[0].caption = "Письмо от {message_sender}\nТема: {subject}\n{plain_text}".format(
                message_sender = msg["from"],
                subject = msg["subject"],
                plain_text = msg["plain_text"]
            )
            for user_id in context.bot_data["enabled_email_distribution"]:
                context.bot.send_media_group(
                    chat_id = user_id,
                    media = attachments
                )
                time.sleep(5)
            for att in msg["attachments_dirs"]:
                os.remove(att)
                os.rmdir("attachments\\"+att.split("\\")[1])
            os.rmdir("attachments")
        else:
            for user_id in context.bot_data["enabled_email_distribution"]:
                context.bot.send_message(
                    chat_id = user_id,
                    text = "Письмо от {message_sender}\nТема: {subject}\n{plain_text}".format(
                        message_sender = msg["from"],
                        subject = msg["subject"],
                        plain_text = msg["plain_text"]
                    )
                )
                time.sleep(5)

def update_schedule_files():
    """Функция для скраппинга файлов расписания с сайта мирэа"""
    try:
        response = requests.get("https://www.mirea.ru/schedule/")
        with open("doc.html", "wb") as htmlfile:
            htmlfile.write(response.content)
        with open("doc.html", "rb") as htmlfile:
            soup = BeautifulSoup(htmlfile, "html.parser")
        os.remove("doc.html")
        div = soup.find("div", {"id": "toggle-hl_2_1-hl_3_3"})
        state = ""
        possible_states = [
            "Расписание экзаменационной сессии:",
            "Расписание занятий:",
            "Расписание зачетной сессии:"
        ]
        filenames = [
            os.path.join("schedule", "exam_2_sem_01_20.xlsx"),
            os.path.join("schedule", "timetable_2_sem_01_20.xlsx"),
            os.path.join("schedule", "assessment_2_sem_01_20.xlsx")
        ]
        for child in div.findChildren("div"):
            if " ".join(child["class"]) == "uk-width-1-1":
                if child.findChildren("b")[0].decode_contents() in possible_states:
                    state = child.findChildren("b")[0].decode_contents()
            elif " ".join(child["class"]) == "uk-width-1-2 uk-width-auto@s":
                if child.findChildren("a")[0].findChildren("div")[0].findChildren("div")[0].decode_contents().find("1 курс") != -1:
                    link_xls = child.findChildren("a")[0]["href"]
                    response = requests.get(link_xls)
                    if not os.path.isdir("schedule"):
                        os.mkdir("schedule")
                    filename = dict(zip(possible_states, filenames))[state]
                    with open(filename, "wb") as xlsfile:
                        xlsfile.write(response.content)
    except:
        print ("Unable to update schedule files")
 
def update_schedule(context: CallbackContext) -> None:
    """Функция для парсинга расписания из .xlsx файлов с сайта миреа, получаемых в update_schedule_files()"""
    
    if os.path.exists(os.path.join("schedule","timetable_2_sem_01_20.xlsx")):
        timetable_wb = openpyxl.load_workbook(os.path.join("schedule","timetable_2_sem_01_20.xlsx"))
        t_ws = timetable_wb.active

        for row in t_ws.iter_rows (min_row=2, max_row=2):
            for cell in row:
                if cell.value == "ИКБО-01-20":
                    group_cell = cell
                    break
    
        week = [
            [[] for i in range (6)],
            [[] for i in range (6)]
        ]

        for col in t_ws.iter_cols (min_row=4, min_col=group_cell.column, max_row=75, max_col=group_cell.column):
            for cell in col:
                subject = {
                    "name": cell.value,
                    "type": t_ws.cell(row=cell.row, column=cell.column+1).value,
                    "classroom": t_ws.cell(row=cell.row, column=cell.column+3).value,
                }
                week[1-(cell.row-4)%2][(cell.row-4)//12].append(subject)
    
        context.bot_data["schedule"] = week
    else:
        print("No schedule file")
    """
    if os.path.exists(os.path.join("schedule","exam_2_sem_01_20.xlsx")):
        exam_wb = openpyxl.load_workbook(os.path.join("schedule", "exam_2_sem_01_20.xlsx")) 
        e_ws = exam_wb.active
    else:
        print("No exam file")
    if os.path.exists(os.path.join("schedule","assessment_2_sem_01_20.xlsx")):
        assessment_wb = openpyxl.load_workbook(os.path.join("schedule", "assessment_2_sem_01_20.xlsx"))
        a_ws = assessment_wb.active
    else:
        print("No assessment file")
    """
    
def get_admins (update: Update) -> list:
    """Функция возвращает список админов"""

    if update.effective_chat.type == "private":
        return []
    return [chat_member.user.id for chat_member in update.effective_chat.get_administrators()]

def check_private_or_not_admin(update: Update) -> bool:
    """Функция проверяет является ли чат приватным или отправитель - админом"""

    if not get_admins(update):
        update.message.reply_text("b-baka")
        return True
    if update.effective_user.id not in get_admins (update):
        update.message.reply_text("Вы не админ!")
        return True
    return False

def enable_schedule_distr(update: Update, context: CallbackContext) -> None:
    if update.effective_user.id != update.effective_chat.id:
        update.message.reply_text("Командой можно воспользоваться только в личных сообщениях!")
        return
    user_id = update.effective_user.id
    try:
        delay = int(context.args[0]) 
    except (IndexError, ValueError):
        update.message.reply_text("Неверное значение минут! Для задержки в 0 минут, напишите /enable_schedule 0")
        return

    if delay<0 or delay>269:
        update.message.reply_text("Неверное значение минут!")
        return
    elif delay==69:
        update.message.reply_text("nice")
        time.sleep(0.5)

    context.bot_data["enabled_schedule_distribution"][user_id]=delay
    reply_text = f"Расписание будет присылаться через {delay} минут после конца последней пары!"
    set_cfg_param("enabled_schedule_distribution", context.bot_data["enabled_schedule_distribution"])
    update.message.reply_text(reply_text)

def disable_schedule_distr(update: Update, context: CallbackContext) -> None:
    if update.effective_user.id != update.effective_chat.id:
        update.message.reply_text("Командой можно воспользоваться только в личных сообщениях!")
        return
    user_id = update.effective_user.id
    if user_id in context.bot_data["enabled_schedule_distribution"].keys():
        context.bot_data["enabled_schedule_distribution"].pop(user_id, None)
        set_cfg_param("enabled_schedule_distribution", context.bot_data["enabled_schedule_distribution"])
    update.message.reply_text ("Пересылка расписания отключена!")

def toggle_email_distr(update: Update, context: CallbackContext) -> None:
    if update.effective_user.id != update.effective_chat.id:
        update.message.reply_text("Командой можно воспользоваться только в личных сообщениях!")
        return
    user_id = update.effective_user.id
    if user_id in context.bot_data["enabled_email_distribution"]:
        context.bot_data["enabled_email_distribution"].remove(user_id)
        reply_text = "Пересылка сообщений с почты отключена!"
    else:
        context.bot_data["enabled_email_distribution"].append(user_id)
        reply_text = "Пересылка сообщений с почты включена!"
    set_cfg_param("enabled_email_distribution", context.bot_data["enabled_email_distribution"])
    update.message.reply_text(reply_text)

def set_cfg_param(param: str, value) -> None:
    if isinstance(value, dict):
        value = ",".join([f"{_key}:{_value}" for _key, _value in value.items()])
    elif isinstance(value, Iterable):
        value = ",".join([str(i) for i in value])
 
    cfgparams = {}
    with open(os.path.join("src", "bot.cfg"), "r") as cfgfile:
        for line in cfgfile:
            line = line[:-1]
            cfgparams[line.split("=")[0]] = line.split("=")[1]

    cfgparams[param] = value

    with open(os.path.join("src", "bot.cfg"), "w") as cfgfile:
        for _key, _value in cfgparams.items():
            cfgfile.write(f"{_key}={_value}\n")

def set_cfg(dispatcher: Dispatcher) -> None:
    #логинимся в почту, получаем id старосты, смотрим список, кто включил рассылки
    email_addr = ""
    email_pass = ""
    imap_host = "imap.gmail.com"
    imap_port = 993
    starosta_id = None
    enabled_schedule_distr = []
    enabled_email_distr = []

    timetable_list = [
        ("09:00", "10:30"),
        ("10:40", "12:10"),
        ("12:40", "14:10"),
        ("14:20", "15:50"),
        ("16:20", "17:50"),
        ("18:00", "19:30"),
    ]
    dispatcher.bot_data["timetable_list"] = timetable_list

    with open(os.path.join("src", "bot.cfg"), "r") as cfgfile:
        for line in cfgfile:
            param, value = line.split("=")
            value = value[:-1]
            if param == "email_address":
                email_addr = value
            if param == "email_password":
                email_pass = value
            if param == "imap_host":
                imap_host = value
            if param == "imap_port":
                imap_port = int(value)
            if param == "starosta_id":
                starosta_id = int(value)
            if param == "enabled_schedule_distribution":
                enabled_schedule_distr = dict(zip([int(i.split(":")[0]) for i in value.split(",")],[int(i.split(":")[1]) for i in value.split(",")]))
            if param == "enabled_email_distribution":
                enabled_email_distr = [int(i) for i in value.split(",")]
            if param == "chat_id":
                dispatcher.bot_data["chat_id"] = int(value)

    try:
        dispatcher.bot_data["emailgetter"] = EmailGetter(
            email_address   = email_addr,
            password  = email_pass,
            imap_host = imap_host,
            imap_port = imap_port
        )
        dispatcher.bot_data["last_email_id"] = dispatcher.bot_data["emailgetter"].get_last_message_id()
        if starosta_id == None:
            print("Не удалось записать id старосты")
        dispatcher.bot_data["starosta_id"] = starosta_id
        dispatcher.bot_data["enabled_schedule_distribution"] = enabled_schedule_distr
        dispatcher.bot_data["enabled_email_distribution"] = enabled_email_distr
    except EmailLoginError:
        print("Не удалось залогиниться в почту")
    #получаем список idшников студентов
    dispatcher.bot_data["students_names_by_ids"] = {}
    with open(os.path.join("src", "ids and names.txt"), "r", encoding="utf-8") as idsfile:
        for line in idsfile:
            line = line[:-1] #удаление \n
            student_id, student_name = int(line.split()[0]), "{0} {1}".format(line.split()[1], line.split()[2])
            dispatcher.bot_data["students_names_by_ids"][student_id] = student_name

    #ставим job_queue на голосование
    tz = pytz.timezone ("Europe/Moscow")     
    t = datetime.time(
        8, 
        00, 
        00, 
        000000, 
        tz
    )

    dispatcher.job_queue.run_daily (
        callback=time_for_poll,
        time=t,
        days=tuple(range(7))
    )

    #ставим job_queue на почту  
    t = datetime.time(
        21, 
        00, 
        00, 
        000000, 
        tz
    )
    dispatcher.job_queue.run_daily(
        callback=get_latest_emails_job_callback,
        time=t,
        days=tuple(range(7)),
        context=dispatcher.bot_data["chat_id"]
    )

    #ставим job_queue на расписание
    t = datetime.time(
        8, 
        00, 
        00, 
        000000, 
        tz
    ) 
    days = (0, 1, 2, 3, 4, 6)

    dispatcher.job_queue.run_daily (
        callback=set_time_table_jobs,
        time=t,
        days=days,
        context=dispatcher.bot_data["chat_id"],
    )

def set_time_table_jobs (context: CallbackContext) -> None:
    
    update_schedule(context)
    
    try:
        parity = week_even()
        day = datetime.datetime.today().weekday()
        day_list = context.bot_data["schedule"][parity][day]
    except (IndexError):
        day_list = []

    time_end = None
    for count, i in enumerate(day_list):
        name = str(i['name'])
        if name == "None":
            continue
        else:
            time_end = context.bot_data["timetable_list"][count][1]

    if time_end == None:
        time_end = "12:00"

    hours, minutes = map(int, time_end.split(":"))
    for user_id in context.bot_data["enabled_schedule_distribution"].keys():
        delayed_time = datetime.datetime.combine(datetime.datetime.now().date(), datetime.time(hours-3, minutes, 00, 000000)) + datetime.timedelta(minutes=context.bot_data["enabled_schedule_distribution"][user_id])
        print(f"ADDING JOB SET_TIME_TABLE FOR USER {user_id}, will be executed on")
        print(context.job_queue.run_once (
            callback=send_time_table,
            when=delayed_time,
            context=user_id
        ).next_t)

def send_time_table (context: CallbackContext) -> None:
    """Отправка сообщения с расписанием"""
    user_id = context.job.context
    next_day(context)

    parity = week_even()
    day = context.bot_data["day_of_week"]
    day_str = context.bot_data["day_of_week_str"]
    day_list = context.bot_data["schedule"][parity][day]
    
    text = ""
    for count, i in enumerate(day_list):
        name = str(i['name'])
        if name == "None":
            continue
        lesson_type = str(i['type']) if i['type'] != None else ""
        classroom = str(i['classroom']) 
        time_beg, time_end= context.bot_data["timetable_list"][count]
        if classroom == "None":
            text+=f"{count+1}) {name} ({lesson_type.upper()}) c {time_beg} до {time_end}\n"
            continue
        elif classroom.upper()== "Д":
            text+=f"{count+1}) {name} ({lesson_type.upper()}) дистанционно c {time_beg} до {time_end}\n"
            continue
        text+=f"{count+1}) {name} ({lesson_type.upper()}) в ауд. {classroom} c {time_beg} до {time_end}\n"
    if text == "":
        text="Завтра пар нет!"
    else:
        text = f"Расписание на {day_str}:\n\n" + text
    context.bot.send_message(
        chat_id=user_id,
        text=text
    )

def week_even () -> int: 
    date = datetime.date.today()
    week = int(date.isocalendar()[1]) - 5
    return week%2

def next_day (context: CallbackContext) -> None:
    day_of_week = datetime.datetime.today().weekday()
    day_of_week_str_ru = {
        "Monday": "понедельник", 
        "Tuesday": "вторник", 
        "Wednesday": "среду", 
        "Thursday": "четверг", 
        "Friday": "пятницу", 
        "Saturday": "субботу"
    }
    if day_of_week==6:
        timetable_day=0
        date = datetime.datetime.today() + datetime.timedelta(days=1)
    else:
        timetable_day=day_of_week+1
        date = datetime.datetime.today() + datetime.timedelta(days=1)
    day_of_week_str= day_of_week_str_ru[datetime.datetime.strptime(f"{date.strftime('%B %d, %Y')}", '%B %d, %Y').strftime('%A')]
    context.bot_data["day_of_week"] = timetable_day
    context.bot_data["day_of_week_str"] = day_of_week_str

def time_for_poll (context: CallbackContext) -> None:
    
    update_schedule(context)
    
    try:
        parity = week_even()
        day = datetime.datetime.today().weekday()
        day_list = context.bot_data["schedule"][parity][day]
    except (IndexError):
        return
    time_beg = None
    for count, i in enumerate(day_list):
        name = str(i['name'])
        if name == "None":
            continue
        else:
            time_beg = context.bot_data["timetable_list"][count][0]
            break
    if time_beg == None:
        return

    hours, minutes = map(int, time_beg.split(":"))    
    poll_time = datetime.datetime.combine(datetime.datetime.now().date(), datetime.time(hours-3, minutes, 00, 000000)) - datetime.timedelta(minutes=30)
    print(f"ADDING JOB POLL, WILL BE EXECUTED AT")
    print(context.job_queue.run_once (
        callback=poll,
        when=poll_time,
        context=context.bot_data["chat_id"]
    ).next_t)

def main() -> None:
    updater = Updater("YOUR BOT TOKEN HERE", use_context=True)
    dispatcher = updater.dispatcher
    set_cfg(dispatcher)
    dispatcher.add_handler(PollAnswerHandler(receive_poll_answer))
    dispatcher.add_handler(CommandHandler('start', helper)) 
    dispatcher.add_handler(CommandHandler('help', helper)) 
    dispatcher.add_handler(CommandHandler('id', id_and_count)) 
    dispatcher.add_handler(CommandHandler("toggle_email", toggle_email_distr)) 
    dispatcher.add_handler(CommandHandler("enable_schedule", enable_schedule_distr))
    dispatcher.add_handler(CommandHandler("disable_schedule", disable_schedule_distr))

    updater.start_polling()
    updater.idle()

if __name__ == '__main__':
    main()
